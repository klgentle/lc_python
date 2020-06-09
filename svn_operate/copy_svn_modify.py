import shutil, os
import openpyxl
import pprint
import codecs
import csv
import os
import time
import platform

# import pysvn

import sys
from sys import argv
from openpyxl import Workbook
from backupToZip import backupToZip
from datetime import datetime

# from send_mail_with_attach import mail
from config_default import configs
from create_date import getBetweenDay
from SvnOperate import SvnOperate
from PlatformOperation import PlatformOperation


class CopyRegister(object):
    """ copy svn/1300_编码 upload register and upload file 
       "usage python[3] copy_upload_ubuntu.py '20190501'"
    """

    def __init__(self, date_str: str):
        self.__date_str_list = [0]
        self.create_date_str_list(date_str)
        print(f"self.__date_str_list:{self.__date_str_list}")

        self.init_path()
        self.svn = SvnOperate(self.__svnup_dir)
        self.svn.update_windows_svn()
        self.make_or_clean_folder()
        self.__data_list = []
        self.__procedure_name_list = []
        self.__error_file_type = set()
        print("init complete")

    def init_path(self):
        home_path = configs.get("path").get("svn_home_path")
        # if self.svn.is_system_windows():
        if platform.uname().system == "Windows":
            home_path = configs.get("path").get("win_svn_home_path")

        if not os.path.exists(home_path):
            home_path = "/mnt/e"

        print(f"home_path:{home_path}")

        self.code_home = os.path.join(home_path, "svn")
        self.__register_folder = os.path.join(self.code_home, "1300_编码", "发布登记表")
        self.__svnup_dir = os.path.join(self.code_home, "1300_编码")

        code_beta_path = os.path.join(home_path, "yx_walk", "beta")
        self.__beta_path = os.path.join(code_beta_path, self.date_str + "beta")

    def create_date_str_list(self, date_str):
        if date_str.find(",") > -1:
            self.date_str = date_str.split(",")[-1]
            self.__date_str_list = getBetweenDay(date_str.split(",")[0], self.date_str)
        else:
            self.date_str = date_str
            self.__date_str_list[0] = self.date_str
        return self.__date_str_list

    def make_or_clean_folder(self):
        if os.path.exists(self.__beta_path):
            print(f"rm -rf {self.__beta_path}")
            shutil.rmtree(f"{self.__beta_path}")
        os.makedirs(self.__beta_path, exist_ok=True)

    def readAllRegister(self):
        """ copy register """
        for folderName, subfolders, filenames in os.walk(self.__register_folder):
            for filename in filenames:
                # find right date register excel
                # filename[-13:-5] is datadate of register
                if filename[-13:-5] not in self.__date_str_list:
                    continue

                whole_filename = os.path.join(folderName, filename)
                self.readOneRegister(whole_filename)

        print(f"data_list count:{len(self.__data_list)}")

    def readOneRegister(self, whole_filename: str):
        """ copy register """
        # copy excel file content
        wb = openpyxl.load_workbook(whole_filename)
        sheet = wb.active
        # skip head
        for row in range(2, sheet.max_row + 1):
            name = sheet["C" + str(row)].value
            # skip blank record
            if not name:
                continue

            # 20 is hard code, if column is max than 20, should change the value
            data_row = [sheet[chr(i + ord("A")) + str(row)].value for i in range(0, 20)]
            path = data_row[4]
            # skip no path record
            if not path:
                print(f"No path, please check register: {data_row}_________________")
                continue
            if isinstance(data_row[7], datetime):
                data_row[7] = data_row[7].strftime("%Y%m%d")

            self.__data_list.append(data_row)
        return self.__data_list

    def get_bo_list(self):
        bo_name_list = []
        for row in self.__data_list:
            name, file_type, path = row[2:5]
            if file_type.strip().upper() in ("RPT", "BO"):
                bo_name_list.append(name.strip())

        return sorted(bo_name_list)


    @staticmethod
    def filename_normlize(filename):
        if filename.find(".") > 0:
            filename = os.path.splitext(filename)[0]
        return filename

    @staticmethod
    def filetype_normlize(file_type):
        """
            > filetype 标准化 
        """
        if not file_type:
            return "sql"
        if file_type.upper() == "BO":
            file_type = "rpt"
        elif file_type.upper() in ("PRO", "FNC", "PRC"):
            file_type = "sql"
        return file_type.lower()

    @staticmethod
    def get_schema_folder(path_list):
        schema_folder = ""
        if path_list[1] == "1370_水晶报表":
            schema_folder = "1370_水晶报表"
        else:
            schema_folder = path_list[2]
        return schema_folder

    def filepath_normlize(self, filepath):
        """
            > filepath 标准化 
        """
        # fix 1300编码 to 1300_编码
        # print(f"filepath:{filepath}________________")
        if filepath.find("1300编码") > -1:
            filepath = filepath.replace("1300编码", "1300_编码")

        # cut path from 1300
        if filepath.find("1300_编码") > -1:
            filepath = filepath[filepath.find("1300_编码") :]
        filepath = PlatformOperation.change_path_sep(filepath)
        return filepath

    def register_data_normalize(self):
        """
            TODO 数据标准化
        """
        file_name_path = map(lambda data_row: data_row[2:5], self.__data_list)
        # print(f"data_list:{self.__data_list}_________________")
        file_name_path = map(
            lambda data: [
                self.filename_normlize(data[0].strip()),
                self.filetype_normlize(data[1].strip()),
                self.filepath_normlize(data[2].strip()),
            ],
            file_name_path,
        )
        return list(file_name_path)

    def register_data_deal(self, file_name_path: list):
        file_name_path_list = [" "] * len(file_name_path)
        for index, data in enumerate(file_name_path):
            name, file_type, path = data

            if file_type.upper() in ("RPT", "BO"):
                name = name.upper()
            file_folder = file_type
            if path.endswith("05Procedures"):
                self.__procedure_name_list.append(name.upper())
                file_folder = "pro"

            whole_file_name = name + "." + file_type
            file_name_path_list[index] = [whole_file_name, file_folder, path]

        return file_name_path_list

    def copyfiles(self, file_name_path_list):
        # copy code files
        for whole_file_name, file_folder, path in file_name_path_list:
            schema_folder = self.get_schema_folder(path.split(os.sep))
            source_file = os.path.join(self.code_home, path, whole_file_name)
            source_file2 = os.path.join(self.code_home, path, whole_file_name.lower())

            target_path = os.path.join(self.__beta_path, schema_folder, file_folder)
            target_path_file = os.path.join(
                target_path, whole_file_name.replace(" ", "_")
            )

            if not os.path.exists(target_path):
                os.makedirs(target_path, exist_ok=True)
            try:
                # copy ignore upper or lower
                shutil.copy(source_file, target_path_file)
            except FileNotFoundError:
                if os.path.exists(source_file2):
                    shutil.copy(source_file2, target_path_file)
                else:
                    self.__error_file_type.add(whole_file_name.split(".")[1])
                    print(f"error! No such file: {source_file} _______________")

        return self.__error_file_type

    def saveRegisterExcel(self):
        "save excel records to one excel"
        file1 = os.path.join(
            self.__svnup_dir, "发布登记表", "cif", "ODS程序版本发布登记表(cif)-template.xlsx"
        )
        file_path_name = self.__beta_path + "/登记表" + self.date_str + ".xlsx"
        shutil.copy(file1, file_path_name)
        # wb = Workbook()
        wb = openpyxl.load_workbook(file_path_name)
        sheet = wb.active

        # TODO remove duplicate, unhashable type: 'list'(error of list(set())) 
        for row in self.__data_list:
            sheet.append(row)
        wb.save(filename=file_path_name)

    def createZipfile(self):
        print(f"file path:{self.__beta_path}")
        return backupToZip(self.__beta_path)

    def list_file2(self, path: str, file_name: str, path2: str):
        to_file = open(file_name, "w")

        to_file.write(f"set define off\nspool {self.date_str}_{path2}.log\n\n")
        for file_name in os.listdir(path):
            # 跳过目标文件
            if file_name in ("pro.sql", "list.sql"):
                continue
            name_without_type = file_name.split(".")[0]
            s = f"prompt\nprompt {name_without_type}\n@@{file_name}\nprompt\nprompt ==============================\nprompt\n"
            to_file.write(s)

        to_file.write("\nspool off\ncommit;\n")
        to_file.close()

    def listSqlFile(self):
        dc = {"pro": "pro.sql", "sql": "list.sql"}
        shema_list = [
            "CBSUSER",
            "ODSUSER",
            "RPTUSER",
            "CBSUSER_MO",
            "ODSUSER_MO",
            "RPTUSER_MO",
        ]
        for schema in shema_list:
            for short_path, file_name in dc.items():
                path = os.path.join(self.__beta_path, schema, short_path)
                whole_file_name = os.path.join(path, file_name)
                # list file
                if os.path.exists(path):
                    self.list_file2(path, whole_file_name, short_path)

    def createConfigCheckSql(self):
        file_name = os.path.join(self.__beta_path, "config_check.sql")
        to_file = open(file_name, "w")
        sql = f"""SELECT OBJECT_NAME FROM ALL_OBJECTS WHERE OWNER = 'RPTUSER' AND OBJECT_TYPE = 'PROCEDURE'
AND OBJECT_NAME IN ({", ".join(["'" + name + "'" for name in self.__procedure_name_list])})
AND SUBSTR(OBJECT_NAME,-4) != '_MID'
MINUS
select OBJECT_NAME from ods_job_config where object_type = 'SP';
        """
        to_file.write(f"{sql}\n")
        to_file.close()

    def write_bo_list(self, file_name="bo_list.txt"):
        bo_name_list = self.get_bo_list()
        print("\n请核对今日BO上线清单：\n" + "\n".join(bo_name_list) + "\n")

        file_name = os.path.join(self.__beta_path, file_name)
        with open(file_name, "w") as to_file:
            to_file.write("请核对今日BO上线清单：\n")
            for b in bo_name_list:
                to_file.write(b + "\n")

    # def send_mail(self, file_path=""):
    #    if not file_path:
    #        file_path = os.path.join(
    #            os.path.basename(self.__beta_path), self.date_str + "beta.zip"
    #        )
    #    mail(self.date_str, file_path)

    def copy_file_from_register(self):
        """main function for call"""
        self.readAllRegister()
        print("read complete")
        self.saveRegisterExcel()
        print("save complete")
        register_data = self.register_data_normalize()
        self.__error_file_type = self.copyfiles(self.register_data_deal(register_data))
        self.listSqlFile()
        print("list file complete")
        self.createConfigCheckSql()
        print("create config done")
        self.write_bo_list()
        print("write bo done")
        self.createZipfile()
        print("all done")

        # if only rpt not find, send email
        # if not self.__error_file_type or self.__error_file_type == {"rpt"}:
        #   a.send_mail()


if __name__ == "__main__":
    # print("usage python[3] copy_upload_ubuntu.py '20190501'")
    # print("usage python[3] copy_upload_ubuntu.py 20181201,20200114")
    date_str = time.strftime("%Y%m%d", time.localtime())
    if len(argv) > 1 and len(argv[1]) == 8:
        if int(date_str) - int(argv[1]) > 10:
            print(f"argv[1] {argv[1]} is too small")
            sys.exit(1)
        date_str = argv[1]
    elif len(argv) > 1:
        date_str = argv[1]

    a = CopyRegister(date_str)
    a.copy_file_from_register()

    print("Done!")
