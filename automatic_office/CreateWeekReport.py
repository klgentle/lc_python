import datetime
import pysnooper
import pprint
import os
import sys
import shutil
import docx
import logging
import openpyxl


# 绝对路径的import
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + "/../")
logging.basicConfig(level=logging.DEBUG, format="%(levelname)s: %(message)s")

from automatic_office.CheckInForm import CheckInForm
from automatic_office.DealInterval import DealInterval


class CreateWeekReport(object):
    """
    根据工作日期，生成周报模板
    """

    def __init__(self, year_month: str):
        if len(year_month) != 6:
            print("Please input year_month with format: YYYYMM!")
            sys.exit(1)
        self.year_month = year_month
        self.interval = DealInterval()
        self.__from_dir = os.path.join(
            os.path.dirname(os.path.abspath(__file__)) + "/../",
            "automatic_office",
            "doc_file",
        )
        self.__target_dir = self.get_target_dir()
        self.from_word = "创兴银行香港ODS项目周报_董坚_fromEndStr.docx"
        self.from_excel = "创兴银行香港ODS项目周报_董坚_fromEndStr.xlsx"

    def get_from_dir(self):
        return self.__from_dir

    def get_target_dir(self):
        target_dir = os.path.join(self.__from_dir, "文思周报", "董坚")
        if not os.path.exists(target_dir):
            os.makedirs(target_dir, exist_ok=True)
        return target_dir

    def copy_file(self, from_file_name, target_file_name: str):
        logging.info("创建文件:%s" % target_file_name)
        shutil.copy(
            os.path.join(self.__from_dir, from_file_name),
            os.path.join(self.__target_dir, target_file_name),
        )

    @staticmethod
    def replace_date_str(content: str, date_name: str, date_str: str) -> str:
        if date_name in content:
            logging.debug("替换文本[%s]为[%s]" % (date_name, date_str))
            content = content.replace(date_name, date_str)
        return content

    def replace_all_str(self, text: str, date_tuple: tuple) -> str:
        text = self.replace_date_str(
            text, "fromEndStr", self.interval.get_from_end_str(date_tuple)
        )
        text = self.replace_date_str(
            text, "fromDate", self.interval.get_from_date(date_tuple)
        )
        text = self.replace_date_str(
            text, "endDate", self.interval.get_end_date(date_tuple)
        )
        return text

    def check_word_change(self, file_name, date_tuple: tuple):
        logging.info("replace word")
        word_file = os.path.join(self.__target_dir, file_name)
        document = docx.Document(word_file)
        for para in document.paragraphs:
            for i, run in enumerate(para.runs):
                run.text = self.replace_all_str(run.text, date_tuple)
        document.save(word_file)

    def check_excel_change(self, file_name, date_tuple: tuple):
        logging.info("replace excel")
        file_path = os.path.join(self.__target_dir, file_name)
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row, col).value in ("fromEndStr", "fromDate", "endDate"):
                    logging.debug("excel value:[%s]" % sheet.cell(row, col).value)
                    # 替换单元格的值
                    sheet.cell(row, col).value = self.replace_all_str(
                        sheet.cell(row, col).value, date_tuple
                    )

        wb.save(file_path)

    def create_week_report(self):
        weekdays = CheckInForm(self.year_month).get_all_work_date()
        logging.info("开始生成周报...")
        for date_tuple in self.interval.get_all_weekday_interval(weekdays, []):
            from_end_str = self.interval.get_from_end_str(date_tuple)

            word_file_name = self.from_word.replace("fromEndStr", from_end_str)
            # copy word file
            self.copy_file(self.from_word, word_file_name)
            # replace content
            self.check_word_change(word_file_name, date_tuple)

            excel_file_name = self.from_excel.replace("fromEndStr", from_end_str)
            # copy excel file
            self.copy_file(self.from_excel, excel_file_name)
            self.check_excel_change(excel_file_name, date_tuple)
        logging.info("周报已生成!")


if __name__ == "__main__":
    if len(sys.argv) <= 1:
        logging.info("Please input year_moth(YYYYMM)")
        sys.exit(1)

    obj = CreateWeekReport(sys.argv[1])
    obj.create_week_report()
