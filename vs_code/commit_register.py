import openpyxl
import shutil, os
import time
from pprint import pprint
from sys import argv

SVN_DIR = "/mnt/e/svn/1300_编码/"
SVN_LOG = "/mnt/e/svn/commit.log"


class Solution:
    """
    """

    def __init__(self, date_str, remark, module_type):
        # svn up
        os.system(f"svn up '{SVN_DIR}'")
        # copy template change excel name
        self.regi_dir = os.path.join(SVN_DIR, "发布登记表", module_type)

        file1 = os.path.join(self.regi_dir, "ODS程序版本发布登记表(支付)-template.xlsx")
        self.date_str = date_str  # str time.strftime("%Y%m%d", time.localtime())
        module_type_name = f"({module_type})"
        if module_type.upper() == "DEPOSIT":
            module_type_name = ""
        self.file2 = os.path.join(
            self.regi_dir, f"ODS程序版本发布登记表{module_type_name}-{self.date_str}.xlsx"
        )
        if not os.path.exists(self.file2):
            shutil.copy(file1, self.file2)

        self.comit_list = []
        self.commit_list_end = ["DONGJIAN", "DEVIL", self.date_str, "", remark]

    def logRead(self):
        svn_log = open(SVN_LOG)
        for line in svn_log.readlines():
            if line.startswith("Sending") or line.startswith("Adding"):
                if line.find("(bin)") > -1:
                    continue

                path_file = line.split("        ")[1]
                path_list = path_file.split("/")
                # 跳过建表语句
                if path_list[-2] == "1380_建表语句":
                    continue

                # print(f"path_list:{path_list}")
                path = "1000_编辑区\\" + "\\".join(path_list[:-1])
                file_list = path_list[-1].split(".")
                # print(f"file_list:{file_list}")
                file_name = file_list[0]
                modu = file_name.upper().split("RPT_")
                module = modu[1][:3]
                print(f"module:{module}")
                file_type = file_list[-1].replace("\n", "")
                row = (
                    [module, "报表"] + [file_name, file_type, path] + self.commit_list_end
                )
                # print(f"row:{row}")
                self.comit_list.append(row)

        svn_log.close()
        # pprint(self.comit_list)

    def logRegister(self):
        wb = openpyxl.load_workbook(self.file2)
        sheet = wb.active

        # DELETE BLANK ROW
        # print(f"max_row{sheet.max_row}")
        if sheet.max_row > 200:
            num = 1
            for row in range(2, sheet.max_row + 1):
                file_name = sheet["C" + str(row)].value
                if file_name:
                    num += 1
                else:
                    # row is blank and row+1 is blank too
                    if not sheet["C" + str(row + 1)].value:
                        break
            print(f"num:{num}")
            sheet.delete_rows(row, sheet.max_row - num)

        for i in self.comit_list:
            sheet.append(i)

        wb.save(filename=self.file2)
        print("excel write down!")


if __name__ == "__main__":
    date_str = time.strftime("%Y%m%d", time.localtime())
    remark = ""
    module_type = "支付"
    # print(f"argv:{argv} ---------- ")
    if len(argv) == 2 and len(argv[1]) == 8:
        date_str = argv[1]
    elif len(argv) == 3:
        date_str, remark = argv[1], argv[2]
    elif len(argv) == 4 and bool(argv[3]):
        date_str, remark, module_type = argv[1], argv[2], argv[3]
    elif len(argv) > 4:
        print("usage: python3 commit_register.py '20190501' remark")
        sys.exit(1)

    a = Solution(date_str, remark, module_type)
    a.logRead()
    a.logRegister()
