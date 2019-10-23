import shutil, os
import openpyxl
import pprint
import codecs
import csv
import os
import time
import platform

import sys
from sys import argv
from openpyxl import Workbook
from backupToZip import backupToZip
from datetime import datetime
from send_mail_with_attach import mail
from config_default import configs
from create_date import getBetweenDay

# from collections import namedtuple


class CopyRegister(object):
    """ copy 'svn/1300_编码' upload register and upload file 
       "usage python[3] copy_upload_ubuntu.py '20190501'"
    """

    def __init__(self, date_str: str):
        self.__date_str_list = [0]
        self.create_date_str_list(date_str)
        print(f"self.__date_str_list:{self.__date_str_list}")

        self.init_path()
        self.update_svn()
        self.make_or_clean_folder()
        self.__data_list = []
        self.__procedure_name_list = []
        self.__error_file_type = set()

    def init_path(self):
        home_path = configs.get("path").get("svn_home_path")
        if self.is_system_windows():
            home_path = configs.get("path").get("win_svn_home_path")

        if not os.path.exists(home_path):
            home_path = "/mnt/e"

        print(f"home_path:{home_path}")

        self.code_home = os.path.join(home_path, "svn")
        self.__register_folder = os.path.join(self.code_home, "1300_编码", "发布登记表")
        self.__svnup_dir = os.path.join(self.code_home, "1300_编码")

        code_beta_path = os.path.join(home_path, "yx_walk", "beta")
        self.__beta_path = os.path.join(code_beta_path, self.date_str + "beta")

    def is_system_windows(self):
        is_system_windows = False
        if platform.uname().system == "Windows":
            is_system_windows = True
        return is_system_windows

    def update_svn(self):
        if self.is_system_windows():
            # BE CAREFUL HERE ###############
            try:
                print("Call svn up ......")
                os.chdir(f"{self.__svnup_dir}")
                os.system("svn up")
            except Exception as e:
                print("SVN UP ERROR: ", e.__doc__)

    def create_date_str_list(self, date_str):
        if date_str.find(",") > -1:
            self.date_str = date_str.split(",")[-1]
            self.__date_str_list = getBetweenDay(date_str.split(",")[0], self.date_str)
        else:
            self.date_str = date_str
            self.__date_str_list[0] = self.date_str
        return self.__date_str_list

    def make_or_clean_folder(self):
        if os.path.exists(self.__beta_path):
            print(f"rm -rf {self.__beta_path}")
            shutil.rmtree(f"{self.__beta_path}")
        os.makedirs(self.__beta_path, exist_ok=True)

    def readAllRegister(self):
        """ copy register """
        for folderName, subfolders, filenames in os.walk(self.__register_folder):
            for filename in filenames:
                # find right date register excel
                # filename[-13:-5] is datadate of register
                if filename[-13:-5] not in self.__date_str_list:
                    continue

                whole_filename = os.path.join(folderName, filename)
                self.readOneRegister(whole_filename)

        print(f"data_list count:{len(self.__data_list)}")

    def readOneRegister(self, whole_filename: str):
        """ copy register """
        # copy excel file content
        wb = openpyxl.load_workbook(whole_filename)
        sheet = wb.active
        for row in range(2, sheet.max_row + 1):
            name = sheet["C" + str(row)].value
            # skip no name row record
            if not name:
                continue
            # 20 is hard code, if column is max than 20, should change the value
            data_row = [sheet[chr(i + ord("A")) + str(row)].value for i in range(0, 20)]
            if isinstance(data_row[7], datetime):
                data_row[7] = data_row[7].strftime("%Y%m%d")

            self.__data_list.append(data_row)
        return self.__data_list

    def register_file_type_deal(self, file_type, path):
        if file_type.upper() == "BO":
            file_type = "rpt"
        elif file_type.upper() in ("PRO", "FNC"):
            file_type = "sql"
        # fixing file_type
        elif file_type.upper() != "RPT" and path.find("1370_水晶报表") > -1:
            file_type = "rpt"
        return file_type

    def get_bo_list(self):
        bo_name_list = []
        for row in self.__data_list:
            name, file_type, path = row[2:5]
            if file_type.upper() in ("RPT", "BO"):
                bo_name_list.append(name)

        return sorted(bo_name_list)

    @staticmethod
    def change_path_sep(path):
        if path.find("\\") > -1 and os.sep != "\\":
            path = path.replace("\\", os.sep)
        elif path.find("/") > -1 and os.sep != "/":
            path = path.replace("/", os.sep)
        return path

    @staticmethod
    def filename_normlize(filename):
        """   > filetype 标准化 """
        return filename.split(".")[0]

    @staticmethod
    def filetype_normlize(filetype):
        if not file_type:
            return "sql"

    def filepath_normlize(self, filepath):
        ind = path.find("1300_编码")
        if ind == -1:
            self.__error_file_type.add("lost")
            print(f"path error, skip path: {path}")
        filepath = change_path_sep(filepath[ind:])
        return filepath

    def register_data_normalize(self):
        file_name_path = map(lambda data_row: data_row[2:5], self.__data_list)
        file_name_path = map(
            lambda data: (
                filename_normlize(data[0]),
                filetype_normlize(data[1]),
                filepath_normlize(data[2]),
            ),
            file_name_path,
        )
        """
            TODO 数据标准化
            > filename 不要加filetype
            > filetype path标准化 
            map
            map
            three functions
            complecat method
        """
        # if file_type.upper() in ("RPT", "BO"):
        #    # name format: rpt to upper
        #    name = name.upper()

        return file_name_path

    def copyfiles(self):
        # copy code files
        for row in self.__data_list:
            name, file_type, path = row[2:5]
            if not file_type:
                file_type = "sql"
            ind = path.find("1300_编码")

            if ind == -1:
                self.__error_file_type.add("lost")
                print(f"path error, skip row: {name}, {file_type}, {path}")
                continue
            file_type = self.register_file_type_deal(file_type, path)

            if file_type.upper() in ("RPT", "BO"):
                # name format: rpt to upper
                name = name.upper()

            # get folder name of code
            # get the file type name to depart pro and sql
            change_path_sep(path)
            path_list = path[ind:].split(os.sep)
            folder_name = path_list[-1]
            # print(f"path_list:{path_list}")
            schema_folder = ""
            if path_list[1] == "1370_水晶报表":
                schema_folder = "1370_水晶报表"
            else:
                schema_folder = path_list[2]

            # folder_name and file type deal
            if folder_name in ("1350_存储过程", "05Procedures"):
                folder_name = "pro"
                self.__procedure_name_list.append(name.upper())
                # file_type in this fold must be sql or pro
                file_type = "sql"
            else:
                folder_name = file_type.lower()

            # strip() delete blank
            name_and_type = name + "." + file_type.lower().strip()
            if name.find(".") > -1:  # too smart
                name_and_type = name

            source_file = os.path.join(self.code_home, path[ind:], name_and_type)
            source_file2 = os.path.join(
                self.code_home, path[ind:], name_and_type.lower()
            )
            # print(f"source_file:{source_file}")

            target_path = os.path.join(self.__beta_path, schema_folder, folder_name)
            # replace blank in file name
            target_path_file = os.path.join(
                target_path, name_and_type.replace(" ", "_")
            )

            if not os.path.exists(target_path):
                os.makedirs(target_path, exist_ok=True)
            try:
                # copy ignore upper or lower
                shutil.copy(source_file, target_path_file)
            except FileNotFoundError:
                if os.path.exists(source_file2):
                    shutil.copy(source_file2, target_path_file)
                else:
                    self.__error_file_type.add(file_type.lower())
                    print(f"error! No such file: {source_file} _______________")

        return self.__error_file_type

    def saveRegisterExcel(self):
        "save excel records to one excel"
        file1 = os.path.join(
            self.__svnup_dir, "发布登记表", "dj", "ODS程序版本发布登记表(dj)-template.xlsx"
        )
        file_path_name = self.__beta_path + "/登记表" + self.date_str + ".xlsx"
        shutil.copy(file1, file_path_name)
        # wb = Workbook()
        wb = openpyxl.load_workbook(file_path_name)
        sheet = wb.active

        # record rows
        for row in self.__data_list:
            sheet.append(row)
        wb.save(filename=file_path_name)

    def createZipfile(self):
        print(f"file path:{self.__beta_path}")
        return backupToZip(self.__beta_path)

    def list_file2(self, path: str, file_name: str, path2: str):
        to_file = open(file_name, "w")

        to_file.write(f"set define off\nspool {self.date_str}_{path2}.log\n\n")
        for file_name in os.listdir(path):
            # 跳过目标文件
            if file_name in ("pro.sql", "list.sql"):
                continue
            name_without_type = file_name.split(".")[0]
            s = f"prompt\nprompt {name_without_type}\n@@{file_name}\nprompt\nprompt ==============================\nprompt\n"
            to_file.write(s)

        to_file.write("\nspool off\ncommit;\n")
        to_file.close()

    def listSqlFile(self):
        dc = {"pro": "pro.sql", "sql": "list.sql"}
        shema_list = [
            "CBSUSER",
            "ODSUSER",
            "RPTUSER",
            "CBSUSER_MO",
            "ODSUSER_MO",
            "RPTUSER_MO",
        ]
        for schema in shema_list:
            for short_path, file_name in dc.items():
                path = os.path.join(self.__beta_path, schema, short_path)
                whole_file_name = os.path.join(path, file_name)
                # list file
                if os.path.exists(path):
                    self.list_file2(path, whole_file_name, short_path)

    def createConfigCheckSql(self):
        file_name = os.path.join(self.__beta_path, "config_check.sql")
        to_file = open(file_name, "w")
        sql = f"""SELECT OBJECT_NAME FROM ALL_OBJECTS WHERE OWNER = 'RPTUSER' AND OBJECT_TYPE = 'PROCEDURE'
AND OBJECT_NAME IN ({", ".join(["'" + name + "'" for name in self.__procedure_name_list])})
AND SUBSTR(OBJECT_NAME,-4) != '_MID'
MINUS
select OBJECT_NAME from ods_job_config where object_type = 'SP';
        """
        to_file.write(f"{sql}\n")
        to_file.close()

    def write_bo_list(self, file_name="bo_list.txt"):
        bo_name_list = self.get_bo_list()
        print("\n请核对今日BO上线清单：\n" + "\n".join(bo_name_list) + "\n")

        file_name = os.path.join(self.__beta_path, file_name)
        with open(file_name, "w") as to_file:
            to_file.write("请核对今日BO上线清单：\n")
            for b in bo_name_list:
                to_file.write(b + "\n")

    def send_mail(self, file_path=""):
        if not file_path:
            file_path = os.path.join(
                os.path.basename(self.__beta_path), self.date_str + "beta.zip"
            )
        mail(self.date_str, file_path)

    def copy_file_from_register(self):
        """main function for call"""
        self.readAllRegister()
        self.saveRegisterExcel()
        self.register_data_normalize()
        self.__error_file_type = self.copyfiles()
        self.listSqlFile()
        self.createConfigCheckSql()
        self.write_bo_list()
        self.createZipfile()

        # if only rpt not find, send email
        # if not self.__error_file_type or self.__error_file_type == {"rpt"}:
        #   a.send_mail()


if __name__ == "__main__":
    # print("usage python[3] copy_upload_ubuntu.py '20190501'")
    date_str = time.strftime("%Y%m%d", time.localtime())
    if len(argv) > 1 and len(argv[1]) == 8:
        if int(date_str) - int(argv[1]) > 10:
            print(f"argv[1] {argv[1]} is too small")
            sys.exit(1)
        date_str = argv[1]
    elif len(argv) > 1:
        date_str = argv[1]

    a = CopyRegister(date_str)
    a.copy_file_from_register()

    print("Done!")
