
import datetime
import openpyxl
import platform
import pysnooper
import sys
import shutil, os
import time

from sys import argv
from date_add import date_add
from copy_svn_modify import CopyRegister

SVN_DIR = "/mnt/e/svn/1300_编码/"
SVN_LOG = "/mnt/e/svn/commit.log"
if platform.uname().system == "Windows":
    SVN_DIR = "E:\\svn\\1300_编码\\"
    SVN_LOG = "E:\\svn\\commit.log"


class Solution:
    """ read log to write excel for install """

    def __init__(self, date_str, mantis, module_type):
        # svn up ########os.system(f"svn up '{SVN_DIR}'")
        # copy template change excel name
        self.regi_dir = os.path.join(SVN_DIR, "发布登记表", module_type)

        file1 = os.path.join(self.regi_dir, "ODS程序版本发布登记表(dj)-template.xlsx")
        self.date_str = date_str  # str time.strftime("%Y%m%d", time.localtime())
        module_type_name = f"({module_type})"
        if module_type.upper() == "DEPOSIT":
            module_type_name = ""
        self.file2 = os.path.join(
            self.regi_dir, f"ODS程序版本发布登记表{module_type_name}-{self.date_str}.xlsx"
        )
        # print(f"self.file2:{self.file2}")
        if not os.path.exists(self.file2):
            shutil.copy(file1, self.file2)

        self.comit_list = []
        self.commit_list_end = ["Dongjian", "Gene", self.date_str, mantis, "", ""]
        # write commit.log
        # os.system("svn st {SVN_DIR} | grep -v '~' > {SVN_LOG}")

    # @pysnooper.snoop()
    def logRead(self):
        svn_log = open(SVN_LOG, encoding="utf-8-sig")  # deal with \ufeff
        for line in svn_log.readlines():
            line = line.strip()
            # print(f"line:{line}")
            if (
                line.startswith("Modified")
                or line.startswith("Modify")
                or line.startswith("Adding")
            ):
                if line.find("ODS程序版本发布登记表") > -1:
                    continue
                # delete  application/octet-stream
                if line.find("application/octet-stream") > -1:
                    line = line.replace("application/octet-stream", "").strip()

                path_file = line[line.find("1300_编码") :]
                path_file = CopyRegister.change_path_sep(path_file)
                path_list = path_file.split(os.sep)
                # 跳过建表语句
                # if path_list[-2] == "1380_建表语句":
                #    continue

                print(f"path_list:{path_list}")
                path = os.path.join("1000_编辑区", os.path.dirname(path_file))
                file_list = path_list[-1].split(".")
                # print(f"file_list:{file_list}")
                file_name = file_list[0]
                modu = file_name.upper().split("RPT_")
                module = ""
                if len(modu) > 1:
                    module = modu[1][:3]

                file_type = file_list[-1].replace("\n", "")
                # 兼容 windows
                path = path.replace(os.sep, "\\")
                row = (
                    [module, "报表"] + [file_name, file_type, path] + self.commit_list_end
                )
                self.comit_list.append(row)

        svn_log.close()

    def logRegister(self):
        wb = openpyxl.load_workbook(self.file2)
        sheet = wb.active

        # DELETE BLANK ROW
        if sheet.max_row > 200:
            num = 1
            for row in range(2, sheet.max_row + 1):
                file_name = sheet["C" + str(row)].value
                if file_name:
                    num += 1
                else:
                    # row is blank and row+1 is blank too
                    if not sheet["C" + str(row + 1)].value:
                        break
            sheet.delete_rows(row, sheet.max_row - num)

        print("records are as bellow:")
        for i in self.comit_list:
            print(f"{i}")
            sheet.append(i)

        wb.save(filename=self.file2)
        print("excel write down!")


if __name__ == "__main__":
    today = time.strftime("%Y%m%d", time.localtime())
    date_str = time.strftime("%Y%m%d", time.localtime())
    time_str = time.strftime("%H:%M", time.localtime())
    if time_str > "16:10":
        # today add one day
        date_str = date_add(1)
    mantis = ""
    module_type = "dj"

    if len(argv) == 2 and len(argv[1]) == 8:
        date_str = argv[1]
    elif len(argv) == 3:
        date_str, mantis = argv[1], argv[2]
    elif len(argv) == 4:
        date_str, mantis, module_type = argv[1], argv[2], argv[3]
        if not argv[3]:
            module_type = "dj"
    elif len(argv) > 4:
        print("usage: python3 commit_register.py '20190501' mantis_id, module_type")
        sys.exit(1)

    if len(argv) > 1 and argv[1].find("d+") > -1:
        # get days from d+days
        days = int(argv[1][2:])
        date_str = date_add(days)

    if date_str < today:
        print(f"date_str:{date_str} is wrong!")
        sys.exit(1)

    print(f"argv:{argv} ---------- ")
    print(f"date_str:{date_str} ---------- ")

    a = Solution(date_str, mantis, module_type)
    a.logRead()
    a.logRegister()
