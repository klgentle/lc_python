import shutil, os
import openpyxl


class CopyExcel(object):
    def __init__(self):
        self.from_path = "/mnt/e/yx_walk/report_develop/report_dept"
        self.from_file = "ods_reports_out.xlsx"
        self.to_file = "ods_reports_out.xlsx"
        #self.sheet_name = sheet_name

    def readData(self):
        wb = openpyxl.load_workbook(os.path.join(self.from_path, self.from_file))
        #sheet = wb.active
        sheets = wb.get_sheet_names()
        print(f"sheets:{sheets}")
        # range from index to change
        for i, v in enumerate(sheets):
            sheet = wb.get_sheet_by_name(v)
            self.data_list = []

            for row in range(1, sheet.max_row + 1):
                name = sheet["A" + str(row)].value
                if not name:
                    continue
                data_row = [sheet[chr(ord("A")+i) + str(row)].value for i in range(0,2)]
                self.data_list.append(data_row)
            # write
            self.saveData()

    def saveData(self):
        file_path_name = os.path.join(self.from_path, self.to_file)
        wb = openpyxl.load_workbook(file_path_name)
        #sheet = wb.active
        #sheet = wb.create_sheet("ALL")
        sheet = wb.get_sheet_by_name("ALL")

        for row in self.data_list:
            #print(f"row:{row}")
            sheet.append(row)

        wb.save(filename=file_path_name)


if __name__ == "__main__":
    #sheet_name = "PMO Report-Recipient Mapping" 
    a = CopyExcel()
    a.readData()
    #a.saveData()
    print(f"All done!")
