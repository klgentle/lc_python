import shutil, os
import openpyxl


class SplitDept(object):
    def __init__(self,sheet_name):
        self.from_path = "/mnt/e/yx_walk/report_develop/report_dept"
        self.from_file = "ods_reports.xlsx"
        self.to_file = "ods_reports_out.xlsx"
        self.data_row = []
        self.data_list = []
        self.sheet_name = sheet_name

    def readData(self):
        wb = openpyxl.load_workbook(os.path.join(self.from_path, self.from_file))
        #sheet = wb.active
        sheet = wb.get_sheet_by_name(self.sheet_name)
        dept_name = ""
        # range from index to change
        for row in range(7, sheet.max_row + 1):
            name = sheet["A" + str(row)].value
            if not name:
                #pass
                continue # continue for blank
            #else:
                #dept_name = name
            #data_row = [sheet[chr(i + ord("A")) + str(row)].value for i in range(0, 2)]
            #dept_name = sheet["B" + str(row)].value
            #dept_name = dept_name.split(' - ')[0]
            #if dept_name.isnumeric():
            #    dept_name = "X"+dept_name
            #data_row = [name, dept_name]
            #print(f"data_row:{data_row}")
            data_row = [sheet["B" + str(row)].value, sheet["E" + str(row)].value]
            self.data_row.append(data_row)

    def saveData(self):
        file_path_name = os.path.join(self.from_path, self.to_file)
        wb = openpyxl.load_workbook(file_path_name)
        #sheet = wb.active
        sheet = wb.create_sheet(self.sheet_name)
        #sheet = wb.get_sheet_by_name(self.sheet_name)

        for row in self.data_list:
            #print(f"row:{row}")
            sheet.append(row)

        wb.save(filename=file_path_name)

    def splitData(self):
        for row in self.data_row:
            rpt_id = row[0]
            rpt_dept = ""

            if len(row) == 1:
                self.data_list.append([rpt_dept, rpt_id])
            else:
                if row[1] and row[1].find(",") > -1:
                    rpt_dept_list = row[1].split(", ") # 中文标点
                    for rpt_dept in rpt_dept_list:
                        self.data_list.append([rpt_dept, rpt_id])
                else:
                    rpt_dept = row[1]
                    self.data_list.append([rpt_dept, rpt_id])


if __name__ == "__main__":
    sheet_name = "PMO Report-Recipient Mapping" 
    a = SplitDept(sheet_name)
    a.readData()
    a.splitData()
    a.saveData()
    print(f"All done!")
