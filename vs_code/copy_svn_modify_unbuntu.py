import shutil, os
import openpyxl
import pprint
import codecs
import csv
import os
import time

# import sh
import sys
from sys import argv
from openpyxl import Workbook
from backupToZip import backupToZip
from datetime import datetime
from send_mail_with_attach import mail
from config_default import configs
from create_date import getBetweenDay

# from list_file import list_file


class CopyRegister(object):
    """ copy 'svn/1300_编码' upload register and upload file 
       "usage python[3] copy_upload_ubuntu.py '20190501'"
    """

    def __init__(self, date_str: str):
        self.__date_str_list = [0]
        # date_str: 20190725,20190730
        self.create_date_str_list(date_str)
        print(f"self.__date_str_list:{self.__date_str_list}")

        home_path = configs.get("path").get("svn_home_path")
        print(f"home_path:{home_path}")
        if not os.path.exists(home_path):
            home_path = "/mnt/e"

        self.code_beta_path = os.path.join(home_path, "yx_walk/beta")
        self.code_home = os.path.join(home_path, "svn")
        self.dir_name = os.path.join(self.code_home, "1300_编码/发布登记表")
        self.svnup_dir = os.path.join(self.code_home, "1300_编码")

        # BE CAREFUL HERE ###############
        # with os.popen("uname -a") as p:
        #    uname = p.read()
        #    if uname.find("Microsoft") == -1:
        #        # linux test
        #        os.system(f"svn up '{self.svnup_dir}'")

        self.__target_path = os.path.join(self.code_beta_path, self.date_str + "beta")
        self.make_or_clean_folder()
        self.__data_list = []
        self.__procedure_name_list = []
        self.__bo_name_list = []

    def create_date_str_list(self, date_str):
        if date_str.find(",") > -1:
            self.date_str = date_str.split(",")[-1]
            self.__date_str_list = getBetweenDay(date_str.split(",")[0], self.date_str)
        else:
            self.date_str = date_str
            self.__date_str_list[0] = self.date_str
        return self.__date_str_list

    def make_or_clean_folder(self):
        if os.path.exists(self.__target_path):
            print(f"rm -rf {self.__target_path}")
            # sh.rm("-rf", f"{self.__target_path}")
            shutil.rmtree(f"{self.__target_path}")
        os.makedirs(self.__target_path, exist_ok=True)

    def readAllRegister(self):
        """ copy register """
        for folderName, subfolders, filenames in os.walk(self.dir_name):
            for filename in filenames:
                # print('FILE INSIDE ' + folderName + ': '+ filename)
                # find right date register excel
                # filename[-13:-5] is datadate of register
                if filename[-13:-5] not in self.__date_str_list:
                    continue
                # test!!!!
                # if filename.find("支付") == -1:
                #    continue

                # print(f"filename: {filename} --------------------------")
                whole_filename = os.path.join(folderName, filename)
                self.readOneRegister(whole_filename)

        # print(f"data_list:{self.__data_list}")
        print(f"data_list count:{len(self.__data_list)}")

    def readOneRegister(self, whole_filename: str):
        """ copy register """
        # copy excel file content
        wb = openpyxl.load_workbook(whole_filename)
        sheet = wb.active
        for row in range(2, sheet.max_row + 1):
            name = sheet["C" + str(row)].value
            # skip no name row record
            if not name:
                continue
            # 20 is hard code, if column is max than 20, should change the value
            data_row = [sheet[chr(i + ord("A")) + str(row)].value for i in range(0, 20)]
            if isinstance(data_row[7], datetime):
                data_row[7] = data_row[7].strftime("%Y%m%d")

            # print(f"data_row:{data_row}")
            self.__data_list.append(data_row)
        return self.__data_list

    def register_file_type_deal(self, file_type, path):
        if file_type.upper() == "BO":
            file_type = "rpt"
        elif file_type.upper() in ("PRO", "FNC"):
            file_type = "sql"
        # fixing file_type
        elif file_type.upper() != "RPT" and path.find("1370_水晶报表") > -1:
            file_type = "rpt"
        return file_type

    def copyfiles(self):
        error_file_type = set()
        # copy code files
        for row in self.__data_list:
            name, file_type, path = row[2:5]
            if not file_type:
                file_type = "sql"
            path = path.replace("\\", "/")
            ind = path.find("1300_编码")

            if ind == -1:
                error_file_type.add("lost")
                print(f"path error, skip row: {name}, {file_type}, {path}")
                continue
            file_type = self.register_file_type_deal(file_type, path)

            if file_type.upper() in ("RPT", "BO"):
                self.__bo_name_list.append(name)
                # name format: rpt to upper
                name = name.upper()

            # get folder name of code
            # get the file type name to depart pro and sql
            path_list = path[ind:].split("/")
            targetName = path_list[-1]
            #print(f"path_list:{path_list}")
            dir_name = ""
            if path_list[1] == "1370_水晶报表":
                dir_name = "1370_水晶报表"
            else:
                dir_name = path_list[2]

            # targetName and file type deal
            if targetName in ("1350_存储过程", "05Procedures"):
                targetName = "pro"
                self.__procedure_name_list.append(name.upper())
                # file_type in this fold must be sql or pro
                file_type = "sql"
            else:
                targetName = file_type.lower()
            # TODO odsuser path and MO path

            # strip() delete blank
            name_and_type = name + "." + file_type.lower().strip()
            if name.find(".") > -1:  # too smart
                name_and_type = name

            source_file = os.path.join(self.code_home, path[ind:], name_and_type)
            source_file2 = os.path.join(
                self.code_home, path[ind:], name_and_type.lower()
            )

            # print(f"source_file:{source_file}")

            target_path2 = os.path.join(self.__target_path, dir_name, targetName)
            if not os.path.exists(target_path2):
                os.makedirs(target_path2, exist_ok=True)
            try:
                # todo copy ignore upper or lower
                shutil.copy(source_file, target_path2)
            except FileNotFoundError:
                if os.path.exists(source_file2):
                    shutil.copy(source_file2, target_path2)
                else:
                    error_file_type.add(file_type.lower())
                    print(f"error! No such file: {source_file} _______________")

        return error_file_type

    def saveRegisterExcel(self):
        "save excel records to one excel"
        file1 = os.path.join(
            self.svnup_dir, "发布登记表", "dj", "ODS程序版本发布登记表(dj)-template.xlsx"
        )
        file_path_name = self.__target_path + "/登记表" + self.date_str + ".xlsx"
        shutil.copy(file1, file_path_name)
        # wb = Workbook()
        wb = openpyxl.load_workbook(file_path_name)
        sheet = wb.active

        # record rows
        for row in self.__data_list:
            sheet.append(row)
        wb.save(filename=file_path_name)

    def createZipfile(self):
        print(f"file path:{self.__target_path}")
        return backupToZip(self.__target_path)

    def list_file(self, path: str, file_name: str, path2: str):
        to_file = open(file_name, "w")
        to_path = f"D:\jdong\\beta\\{self.date_str}beta\\{path2}"
        for f in os.listdir(path):
            s = f"@@{to_path}\{f};\n"
            to_file.write(s)

        to_file.write("commit;\n")
        to_file.close()

    def list_file2(self, path: str, file_name: str, path2: str):
        to_file = open(file_name, "w")

        to_file.write(f"set define off\nspool {self.date_str}_{path2}.log\n\n")
        for file_name in os.listdir(path):
            # 跳过目标文件
            if file_name in ("pro.sql", "list.sql"):
                continue
            name_without_type = file_name.split(".")[0]
            s = f"prompt\nprompt {name_without_type}\n@@{file_name}\nprompt\nprompt ==============================\nprompt\n"
            to_file.write(s)

        to_file.write("\nspool off\ncommit;\n")
        to_file.close()

    def listSqlFile(self):
        dc = {"pro": "pro.sql", "sql": "list.sql"}
        shema_list = [
            "CBSUSER",
            "ODSUSER",
            "RPTUSER",
            "CBSUSER_MO",
            "ODSUSER_MO",
            "RPTUSER_MO",
        ]
        for schema in shema_list:
            for short_path, file_name in dc.items():
                path = os.path.join(self.__target_path, schema, short_path)
                whole_file_name = os.path.join(path, file_name)
                # list file
                if os.path.exists(path):
                    self.list_file2(path, whole_file_name, short_path)

    def createConfigCheckSql(self):
        file_name = os.path.join(self.__target_path, "config_check.sql")
        to_file = open(file_name, "w")
        # print test
        # print(f"self.__procedure_name_list:{self.__procedure_name_list}")
        sql = f"""SELECT OBJECT_NAME FROM ALL_OBJECTS WHERE OWNER = 'RPTUSER' AND OBJECT_TYPE = 'PROCEDURE'
AND OBJECT_NAME IN ({", ".join(["'" + name + "'" for name in self.__procedure_name_list])})
AND SUBSTR(OBJECT_NAME,-4) != '_MID'
MINUS
select OBJECT_NAME from ods_job_config where object_type = 'SP';
        """
        to_file.write(f"{sql}\n")
        to_file.close()

    def boNameList(self, file_name="bo_list.txt"):
        self.__bo_name_list.sort()
        print("\n请核对今日BO上线清单：\n" + "\n".join(self.__bo_name_list) + "\n")

        file_name = os.path.join(self.__target_path, file_name)
        to_file = open(file_name, "w")
        to_file.write("请核对今日BO上线清单：\n")
        for b in self.__bo_name_list:
            to_file.write(b + "\n")

        to_file.close()

    def send_mail(self, file_path=""):
        if not file_path:
            file_path = os.path.join(self.code_beta_path, self.date_str + "beta.zip")
        mail(self.date_str, file_path)


if __name__ == "__main__":
    date_str = time.strftime("%Y%m%d", time.localtime())
    if len(argv) > 1 and len(argv[1]) == 8:
        if int(date_str) - int(argv[1]) > 10:
            print(f"argv[1] {argv[1]} is to small")
            sys.exit(1)
        if int(date_str) < int(argv[1]):
            print(f"argv[1] {argv[1]} is large than today")
            sys.exit(1)
        date_str = argv[1]
    elif len(argv) > 1:
        date_str = argv[1]

    a = CopyRegister(date_str)
    a.readAllRegister()
    a.saveRegisterExcel()
    error_file_type = a.copyfiles()
    a.listSqlFile()
    a.createConfigCheckSql()
    a.boNameList()
    # not create zip file, need to add rpt files
    a.createZipfile()

    # if only rpt not find, send email
    if not error_file_type or error_file_type == {"rpt"}:
        a.send_mail()

    print("Done!")
    # print("usage python[3] copy_upload_ubuntu.py '20190501'")
